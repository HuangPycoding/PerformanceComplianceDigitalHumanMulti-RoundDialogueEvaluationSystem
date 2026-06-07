"""合并所有生成的评测用例，履约场景检查，写入 xlsx 和 txt"""
import json
import openpyxl

# 读取所有批次
all_cases = []
for fname in ["generated_cases.json", "generated_cases_batch2.json", "generated_cases_batch3.json", "generated_cases_batch4.json"]:
    with open(fname, "r", encoding="utf-8") as f:
        cases = json.load(f)
        all_cases.extend(cases)
        print(f"读取 {fname}: {len(cases)} 条")

print(f"合计新增: {len(all_cases)} 条")

# ========== 履约场景检查 ==========
print("\n" + "=" * 60)
print("履约数字人场景贴合度检查")
print("=" * 60)

# 履约核心要素关键词
fulfillment_kw = [
    "配送", "订单", "履约", "合同", "骑手", "送达", "提货", "入住",
    "退款", "改签", "预约", "还款", "到期", "过期",
    "售后", "出餐", "接单", "取餐", "闭园", "延迟", "异常",
    "扣费", "归还", "计费", "关锁", "投诉", "回访",
    "通知", "提醒", "确认", "核实", "替换", "缺货",
    "变更", "开通", "上线", "报名", "奖励", "激励",
    "安全", "赔付", "补偿", "减免", "考核", "改期",
    "暂停", "恢复", "申诉", "预警", "达标",
]

all_pass = True
for case in all_cases:
    inst = case["instruction"]
    score = sum(1 for kw in fulfillment_kw if kw in inst)
    status = "PASS" if score >= 3 else "REVIEW"
    if status == "REVIEW":
        all_pass = False
    print(f"  #{case['id']:2d} {case['title']}")
    print(f"      履约关键词命中 {score} 个 → {status}")

print(f"\n{'全部通过履约场景检查' if all_pass else '存在需复核的用例，请检查'}")
print("=" * 60)

# 读取原 xlsx
src = "命题二：外呼任务对话模型指令示例.xlsx"
wb = openpyxl.load_workbook(src)
ws = wb.active

# 写入新数据
for case in all_cases:
    ws.cell(row=case["id"] + 1, column=1, value=case["id"])
    ws.cell(row=case["id"] + 1, column=2, value=case["instruction"])

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 80

# 保存 xlsx
xlsx_out = "外呼任务对话模型指令示例_v4_60条.xlsx"
wb.save(xlsx_out)
print(f"\nxlsx 已保存: {xlsx_out}")

# 保存 txt
txt_out = "外呼任务对话模型指令示例_v4_60条.txt"
with open(txt_out, "w", encoding="utf-8") as f:
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                f.write(str(cell) + "\n")
            f.write("===CELL_END===\n")
print(f"txt 已保存: {txt_out}")

print(f"\n总计: {ws.max_row - 1} 条用例（原有2条 + 新增{len(all_cases)}条）")
print("\n完整索引 (60条):")
# Original 2
original = [(1, "美团外卖 - 骑手飞毛腿合同通知"), (2, "课程平台 - 直播选项升级通知")]
for cid, title in original:
    print(f"  #{cid:2d}  {title} (原有)")
for c in all_cases:
    print(f"  #{c['id']:2d}  {c['title']}")
# Count by batch
b1 = sum(1 for c in all_cases if 3 <= c['id'] <= 9)
b2 = sum(1 for c in all_cases if 10 <= c['id'] <= 17)
b3 = sum(1 for c in all_cases if 18 <= c['id'] <= 20)
b4 = sum(1 for c in all_cases if 21 <= c['id'] <= 30)
print(f"\n批次分布: 原有2 + 批次1({b1}) + 批次2({b2}) + 批次3({b3}) + 批次4({b4}) = 30")
